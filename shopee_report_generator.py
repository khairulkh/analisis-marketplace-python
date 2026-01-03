"""
Generator laporan Excel untuk analisis Shopee
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ShopeeReportGenerator:
    """Class untuk generate laporan Excel Shopee"""
    
    def __init__(self):
        # Define colors for Excel
        self.colors = {
            'header': '366092',  # Dark blue
            'subheader': '4F81BD',  # Medium blue
            'high_priority': 'C00000',  # Red
            'medium_priority': 'FFC000',  # Yellow
            'low_priority': '00B050',  # Green
            'boncos': 'FF0000',  # Bright red
            'rugi': 'FFC7CE',  # Light red
            'break_even': 'FFEB9C',  # Light yellow
            'untung': 'C6EFCE',  # Light green
            'untung_tinggi': '92D050'  # Bright green
        }
        
        # Define styles
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.subheader_font = Font(name='Calibri', size=11, bold=True)
        self.normal_font = Font(name='Calibri', size=10)
        self.bold_font = Font(name='Calibri', size=10, bold=True)
        
    def generate_excel_report(self, raw_data, cleaned_data, analysis_results, 
                             campaign_summary, daily_summary, file_name=None):
        """Generate comprehensive Excel report"""
        if file_name is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"shopee_analysis_report_{timestamp}.xlsx"
        
        print(f"\n💾 Generating Excel report: {file_name}")
        
        # Create Excel writer
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            # Sheet 1: Raw Data
            raw_data.to_excel(writer, sheet_name='Raw_Data', index=False)
            
            # Sheet 2: Cleaned Data
            cleaned_data.to_excel(writer, sheet_name='Cleaned_Data', index=False)
            
            # Sheet 3: Campaign Analysis
            analysis_results.to_excel(writer, sheet_name='Campaign_Analysis', index=False)
            
            # Sheet 4: Campaign Summary
            campaign_summary.to_excel(writer, sheet_name='Campaign_Summary', index=False)
            
            # Sheet 5: Daily Summary
            if not daily_summary.empty:
                daily_summary.to_excel(writer, sheet_name='Daily_Summary', index=False)
            
            # Sheet 6: Recommendations
            self._create_recommendations_sheet(writer, analysis_results)
            
            # Sheet 7: Strategy Guide
            self._create_strategy_guide_sheet(writer)
            
            # Sheet 8: Performance Dashboard
            self._create_dashboard_sheet(writer, analysis_results, campaign_summary)
        
        # Apply formatting
        self._apply_excel_formatting(file_name)
        
        print(f"✅ Excel report generated: {file_name}")
        return file_name
    
    def _create_recommendations_sheet(self, writer, analysis_results):
        """Create recommendations sheet"""
        recommendations_data = []
        
        # Sort by priority
        priority_order = {'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        analysis_results['Priority_Num'] = analysis_results['Priority'].map(priority_order)
        sorted_results = analysis_results.sort_values('Priority_Num')
        
        for _, row in sorted_results.iterrows():
            recommendations_data.append({
                'Priority': row['Priority'],
                'Campaign': row['Campaign'],
                'Status': row['Status'],
                'ROAS': row['ROAS'],
                'Performance_Score': row['Performance_Score'],
                'Action_Required': row['Recommendations'],
                'Budget_Advice': row['Budget_Advice'],
                'Focus_Area': row['Focus_Area'],
                'Timeline': self._get_timeline_by_priority(row['Priority'])
            })
        
        rec_df = pd.DataFrame(recommendations_data)
        rec_df.to_excel(writer, sheet_name='Recommendations', index=False)
    
    def _create_strategy_guide_sheet(self, writer):
        """Create strategy guide sheet"""
        # Daily schedule
        daily_schedule = pd.DataFrame({
            'Time': ['07:00-09:00', '12:00-13:00', '20:00-22:00', '00:00-01:00'],
            'Activity': [
                'Adjust ROAS/budget if needed\nShopee starts pushing ads',
                'Monitor CTR & temporary clicks\nCheck budget absorption',
                'Complete daily performance evaluation\nCompare with targets',
                'DO NOT CHANGE ANYTHING!\nShopee learning reset phase'
            ],
            'Priority': ['HIGH', 'MEDIUM', 'HIGH', 'LOW'],
            'Can_Change_Settings': ['YES', 'NO', 'YES', 'NO']
        })
        
        # Weekly rhythm
        weekly_rhythm = pd.DataFrame({
            'Day': ['1 (Monday)', '2 (Tuesday)', '3 (Wednesday)', '4 (Thursday)', 
                   '5 (Friday)', '6 (Saturday)', '7 (Sunday)'],
            'Focus': [
                'Evaluate weekend data\nAnalyze trends',
                'Monitor consistency\nDO NOT change settings',
                'Execute scale up/down\nIncrease budget 20-30%',
                'Let system learn\nDO NOT change anything',
                'Evaluate scale results\nRollback if dropped',
                'Maintenance & split testing\nTest new creatives',
                'Weekly review & planning\nDuplicate successful campaigns'
            ],
            'Action': ['EVALUATE', 'MONITOR', 'EXECUTE', 'OBSERVE', 
                      'EVALUATE', 'TEST', 'PLAN']
        })
        
        # Scale criteria
        scale_criteria = pd.DataFrame({
            'Criteria': [
                'Sales increase ≥10-20% for 2 consecutive days',
                'Revenue increase ≥15-25%',
                'Clicks increase ≥20%',
                'CTR stable >2.5%',
                'Actual ROAS > target ROAS setting',
                'Daily budget absorption >70%'
            ],
            'Metric': ['Produk Terjual', 'Omzet Penjualan', 'Jumlah Klik', 
                      'Persentase Klik', 'Efektifitas Iklan', 'Budget Usage'],
            'Target': ['10-20% ↑', '15-25% ↑', '20% ↑', '>2.5%', '>Target', '>70%']
        })
        
        # Write to Excel
        daily_schedule.to_excel(writer, sheet_name='Strategy_Guide', 
                               startrow=0, index=False)
        weekly_rhythm.to_excel(writer, sheet_name='Strategy_Guide', 
                              startrow=len(daily_schedule) + 3, index=False)
        scale_criteria.to_excel(writer, sheet_name='Strategy_Guide', 
                               startrow=len(daily_schedule) + len(weekly_rhythm) + 6, 
                               index=False)
    
    def _create_dashboard_sheet(self, writer, analysis_results, campaign_summary):
        """Create performance dashboard sheet"""
        dashboard_data = []
        
        # Overall metrics
        total_campaigns = len(campaign_summary)
        total_spend = campaign_summary['Spend'].sum()
        total_sales = campaign_summary['Sales'].sum()
        total_profit = campaign_summary['Profit'].sum()
        overall_roas = total_sales / total_spend if total_spend > 0 else 0
        
        # Campaign status distribution
        status_counts = analysis_results['Status'].value_counts()
        
        # Top performers
        top_3 = campaign_summary.nlargest(3, 'ROAS')
        bottom_3 = campaign_summary.nsmallest(3, 'ROAS')
        
        # Prepare dashboard data
        dashboard_data.append(['OVERALL PERFORMANCE METRICS', '', ''])
        dashboard_data.append(['Total Campaigns', total_campaigns, ''])
        dashboard_data.append(['Total Spend', f"Rp {total_spend:,.0f}", ''])
        dashboard_data.append(['Total Sales', f"Rp {total_sales:,.0f}", ''])
        dashboard_data.append(['Total Profit', f"Rp {total_profit:,.0f}", 
                             'Green if profit, Red if loss'])
        dashboard_data.append(['Overall ROAS', f"{overall_roas:.2f}", 
                             f"{'✅ Excellent' if overall_roas > 1.5 else '⚠️ Needs attention'}"])
        dashboard_data.append(['', '', ''])
        
        dashboard_data.append(['CAMPAIGN STATUS DISTRIBUTION', 'Count', 'Percentage'])
        for status, count in status_counts.items():
            percentage = (count / total_campaigns) * 100
            dashboard_data.append([status, count, f"{percentage:.1f}%"])
        dashboard_data.append(['', '', ''])
        
        dashboard_data.append(['TOP 3 PERFORMERS (by ROAS)', 'ROAS', 'Profit'])
        for _, row in top_3.iterrows():
            dashboard_data.append([
                row['Campaign'][:30] + ('...' if len(row['Campaign']) > 30 else ''),
                f"{row['ROAS']:.2f}",
                f"Rp {row['Profit']:,.0f}"
            ])
        dashboard_data.append(['', '', ''])
        
        dashboard_data.append(['NEEDS ATTENTION (ROAS < 1)', 'ROAS', 'Loss'])
        for _, row in bottom_3.iterrows():
            if row['ROAS'] < 1:
                dashboard_data.append([
                    row['Campaign'][:30] + ('...' if len(row['Campaign']) > 30 else ''),
                    f"{row['ROAS']:.2f}",
                    f"Rp {row['Spend'] - row['Sales']:,.0f}"
                ])
        
        # Convert to DataFrame
        dashboard_df = pd.DataFrame(dashboard_data)
        dashboard_df.to_excel(writer, sheet_name='Performance_Dashboard', 
                             index=False, header=False)
    
    def _get_timeline_by_priority(self, priority):
        """Get timeline based on priority"""
        timelines = {
            'HIGH': 'IMMEDIATE (Today)',
            'MEDIUM': 'WITHIN 2-3 DAYS',
            'LOW': 'THIS WEEK'
        }
        return timelines.get(priority, 'WHEN POSSIBLE')
    
    def _apply_excel_formatting(self, file_path):
        """Apply formatting to Excel file"""
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format headers
                self._format_headers(ws, sheet_name)
                
                # Format cells based on values
                self._format_cells(ws, sheet_name)
                
                # Auto-adjust column widths
                self._auto_adjust_columns(ws)
            
            wb.save(file_path)
            print("✅ Excel formatting applied")
            
        except Exception as e:
            print(f"⚠️ Could not apply Excel formatting: {e}")
    
    def _format_headers(self, ws, sheet_name):
        """Format header row"""
        if ws.max_row > 0 and ws.max_column > 0:
            header_fill = PatternFill(start_color=self.colors['header'],
                                     end_color=self.colors['header'],
                                     fill_type='solid')
            
            for cell in ws[1]:
                cell.font = self.header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _format_cells(self, ws, sheet_name):
        """Format cells based on values"""
        if sheet_name == 'Campaign_Analysis':
            # Find ROAS column
            roas_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == 'ROAS':
                    roas_col = col
                    break
            
            if roas_col:
                for row in range(2, ws.max_row + 1):
                    roas_value = ws.cell(row=row, column=roas_col).value
                    if isinstance(roas_value, (int, float)):
                        cell = ws.cell(row=row, column=roas_col)
                        
                        if roas_value >= 2:
                            cell.fill = PatternFill(start_color=self.colors['untung_tinggi'],
                                                   end_color=self.colors['untung_tinggi'],
                                                   fill_type='solid')
                        elif roas_value >= 1.2:
                            cell.fill = PatternFill(start_color=self.colors['untung'],
                                                   end_color=self.colors['untung'],
                                                   fill_type='solid')
                        elif roas_value >= 1:
                            cell.fill = PatternFill(start_color=self.colors['break_even'],
                                                   end_color=self.colors['break_even'],
                                                   fill_type='solid')
                        elif roas_value >= 0.5:
                            cell.fill = PatternFill(start_color=self.colors['rugi'],
                                                   end_color=self.colors['rugi'],
                                                   fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color=self.colors['boncos'],
                                                   end_color=self.colors['boncos'],
                                                   fill_type='solid')
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width